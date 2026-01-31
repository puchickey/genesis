
import openpyxl
import sys
import os

def extract_text_from_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_content.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                # Filter out completely empty rows to reduce noise
                if any(row_text):
                    text_content.append("\t".join(row_text))
        return "\n".join(text_content)
    except Exception as e:
        return f"Error reading {file_path}: {str(e)}"

if __name__ == "__main__":
    target_file = r"G:\マイドライブ\Genesis_OS\02_仕事\01_投資活動\転職活動\履歴書_20251125.xlsx"
    if os.path.exists(target_file):
        print(extract_text_from_xlsx(target_file))
    else:
        print(f"File not found: {target_file}")
